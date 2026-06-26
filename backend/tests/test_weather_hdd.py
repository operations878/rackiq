"""Tests for the HDD / weather Data Studio source (Stage 0 ingestion).

The HDD book becomes a re-uploadable, idempotent source. The parser finds the header + date/year axis
empirically and handles BOTH a tidy real-date layout and a by-year matrix. HDD ≡ max(0, 65 − tmean).
"""

from __future__ import annotations

import datetime as dt
import tempfile

import numpy as np
from openpyxl import Workbook

from app import db, weather_hdd


def _seasonal_hdd(d: dt.date) -> float:
    doy = d.timetuple().tm_yday
    t = 50.0 - 18.0 * np.cos(2 * np.pi * (doy - 15) / 365.0)
    return max(0.0, 65.0 - t)


def _make_tidy(path):
    """Real dates × named HDD / Normal / 5-Yr / 10-Yr / BX HO SOLD columns, with title noise above."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HDD'S"
    ws.append(["LaGuardia Heating Degree Days", None, None, None, None, None])
    ws.append([None, None, None, None, None, None])
    ws.append(["Date", "HDD", "Normal", "5-Yr Avg", "10-Yr Avg", "BX HO SOLD"])
    d = dt.date(2022, 1, 1)
    rng = np.random.default_rng(1)
    for _ in range(365 * 2):
        h = _seasonal_hdd(d)
        ws.append([d, round(h, 1), round(h, 1), round(h * 1.02, 1), round(h * 0.98, 1),
                   round(max(0.0, 1000 + 120 * h + rng.normal(0, 300)), 0)])
        d += dt.timedelta(days=1)
    wb.save(path)


def _make_matrix(path):
    """Month/day rows × year columns (+ a Normal baseline column)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HDD's"
    ws.append(["Bronx HDD by year", None, None, None, None])
    ws.append(["Day", 2020, 2021, 2022, "Normal"])
    base = dt.date(2020, 7, 1)
    for i in range(365):
        d = base + dt.timedelta(days=i)
        vals = [round(_seasonal_hdd(dt.date(y, d.month, d.day)), 1) for y in (2020, 2021, 2022)]
        ws.append([f"{d.month}/{d.day}"] + vals + [round(_seasonal_hdd(d), 1)])
    wb.save(path)


def test_tidy_layout_parses_with_anchor(con):
    path = tempfile.mktemp(suffix=".xlsx")
    _make_tidy(path)
    res = weather_hdd.load_hdd_file(con, path, "now")
    diag = res["diagnostics"]
    assert diag["mode"] == "tidy"
    assert res["observations_written"] == 730
    assert res["station"] == "LGA"
    assert {"hdd", "hdd_normal", "hdd_5yr", "hdd_10yr", "ho_sold"} <= set(diag["roles_mapped"])
    # the BX HO SOLD anchor is date-alignable in a tidy layout → monthly pairs landed
    assert res["anchor_months_written"] == 24
    anc = weather_hdd.read_anchor(con, "LGA")
    assert len(anc) == 24 and (anc["ho_sold"] > 0).all()
    assert (anc["hdd_month"] >= 0).all() and (anc["hdd_month"] > 0).any()   # summer months are 0 HDD


def test_year_matrix_layout_melts_to_dated_observations(con):
    path = tempfile.mktemp(suffix=".xlsx")
    _make_matrix(path)
    res = weather_hdd.load_hdd_file(con, path, "now")
    diag = res["diagnostics"]
    assert diag["mode"] == "year_matrix"
    assert diag["year_columns"] == [2020, 2021, 2022]
    assert res["observations_written"] == 365 * 3
    hd = weather_hdd.read_hdd(con, "LGA")
    assert hd["day"].min().year == 2020 and hd["day"].max().year == 2022


def test_reupload_is_idempotent(con):
    path = tempfile.mktemp(suffix=".xlsx")
    _make_tidy(path)
    weather_hdd.load_hdd_file(con, path, "now")
    before = weather_hdd.store_counts(con)["hdd_observations"]
    weather_hdd.load_hdd_file(con, path, "later")     # re-upload the same file
    after = weather_hdd.store_counts(con)["hdd_observations"]
    assert before == after == 730                      # upsert on (station, day) — no double count


def test_hdd_identity_verified_against_mean_temp(con):
    """When a mean-temp column is present, HDD = max(0, 65 − tmean) is verified (mismatches reported)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HDD'S"
    ws.append(["Date", "Mean Temp", "HDD"])
    # row 1 consistent (tmean 40 → HDD 25); row 2 deliberately wrong (tmean 40 → HDD 99)
    ws.append([dt.date(2023, 1, 1), 40.0, 25.0])
    ws.append([dt.date(2023, 1, 2), 40.0, 99.0])
    for i in range(3, 30):
        t = 30.0 + i
        ws.append([dt.date(2023, 1, i), t, round(max(0.0, 65 - t), 1)])
    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    res = weather_hdd.load_hdd_file(con, path, "now")
    assert res["diagnostics"]["identity_checked"] >= 20
    assert res["diagnostics"]["identity_mismatches"] == 1   # only the bad row


def test_station_inference():
    assert weather_hdd.infer_station("LaGuardia HDD report") == "LGA"
    assert weather_hdd.infer_station("Newark degree days") == "EWR"
    assert weather_hdd.infer_station("Baltimore") == "BWI"
    assert weather_hdd.infer_station("something else") == "LGA"   # default = the LGA book
