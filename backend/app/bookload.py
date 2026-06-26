"""Repeatable real-book loaders — Account Reference Chart, raw BOLs, and the deal book.

These turn the operator's raw files (dropped into ``sample_data/deals/``) into the canonical store,
the same way every run, so re-loading an updated book is one call (CLI ``rackiq-load-realbook`` or
the Data Studio "Deals"/"BOLs" sources):

  • Account Reference Chart (raw BOL account name → coded master) → the customer crosswalk.
  • Raw BOLs (compartments) → ``lifts`` (group by BOL, sum gross+net, drop 0/0/0 control rows, use
    Ship Date, product → family), with the consignee name resolved to its coded master.
  • Term + forward-fixed + spot → the ``deals`` table (idempotent), masters resolved via the bridge.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from . import crosswalk, db, dealbook, schema
from .ingest import _norm


# ---- Account Reference Chart (raw → coded) → crosswalk ---------------------------
def parse_account_chart(path: str) -> list[tuple[str, str]]:
    """Read the 2-column chart (Raw BOL Account Names → Coded Account Names)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    pairs: list[tuple[str, str]] = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:                       # header row
            continue
        raw = r[0] if len(r) > 0 else None
        coded = r[1] if len(r) > 1 else None
        if raw is not None and coded is not None and str(raw).strip() and str(coded).strip():
            pairs.append((str(raw).strip(), str(coded).strip()))
    wb.close()
    return pairs


def load_account_chart(con, path: str, now: str) -> dict:
    """Load the chart as CONFIRMED crosswalk entries (raw → coded master). Returns load stats."""
    pairs = parse_account_chart(path)
    out = crosswalk.load_name_map(con, pairs, now)
    out["pairs"] = pairs
    return out


# ---- Raw BOLs → lifts ------------------------------------------------------------
_BOL_USECOLS = ["Consignee Number", "Consignee Name", "Ship Date", "BOL Number", "Terminal Name",
                "Terminal Product", "Product Name", "Gross Amount", "Net Amount", "Temperature",
                "Gravity (API)"]


def load_bols(con, paths, chart_pairs: list[tuple[str, str]] | None, now: str,
              replace: bool = True) -> dict:
    """Load raw BOL compartments into ``lifts`` (one lift per BOL). Accepts one path or several
    (e.g. multiple years — Jul 2023→Jun 2024 plus a later book — which are concatenated).

    Drops 0/0/0 control rows (BOL=0 & gross=0 & net=0), groups compartments sharing a BOL into one
    lift (gross + net summed), uses **Ship Date** (not Submission Date), maps the product to a family,
    and resolves the raw consignee NAME to its coded master via the chart (normalized match — so
    trivial whitespace/case differences don't drop a customer). Unmapped consignees keep their raw
    name and surface in the unmapped panel.
    """
    if isinstance(paths, str):
        paths = [paths]
    frames = []
    for p in paths:
        f = pd.read_csv(p, dtype=object, usecols=lambda c: c.strip() in _BOL_USECOLS)
        f.columns = [c.strip() for c in f.columns]
        frames.append(f)
    raw = pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]
    for c in ("Consignee Name", "Terminal Name", "Product Name", "Consignee Number", "BOL Number"):
        if c in raw.columns:
            raw[c] = raw[c].map(lambda v: str(v).strip() if v is not None and not pd.isna(v) else None)
    gross = pd.to_numeric(raw["Gross Amount"], errors="coerce")
    net = pd.to_numeric(raw["Net Amount"], errors="coerce")
    bol_n = pd.to_numeric(raw["BOL Number"], errors="coerce")
    temp = pd.to_numeric(raw["Temperature"], errors="coerce")
    api = pd.to_numeric(raw["Gravity (API)"], errors="coerce")
    ship = pd.to_datetime(raw["Ship Date"], errors="coerce")

    # drop EDI control / heartbeat rows: BOL=0 & gross=0 & net=0
    control = (bol_n.fillna(-1) == 0) & (gross.fillna(0) == 0) & (net.fillna(0) == 0)
    keep = ~control & ship.notna() & (net.notna() | gross.notna())
    df = pd.DataFrame({
        "consignee_name": raw["Consignee Name"],
        "consignee_number": raw["Consignee Number"],
        "bol_number": raw["BOL Number"],
        "ship_date": ship,
        "terminal": raw["Terminal Name"],
        "product_raw": raw["Product Name"],
        "gross": gross, "net": net, "temp": temp, "api": api,
    })[keep].copy()
    n_control = int(control.sum())
    n_compartments = int(len(df))

    df["product"] = df["product_raw"].map(dealbook.product_family)
    # one lift per (BOL, ship date, consignee) — guards against BOL-number reuse across days/customers
    grp = df.groupby(["bol_number", "ship_date", "consignee_number"], dropna=False)
    lifts = grp.agg(
        net_gallons=("net", "sum"),
        gross_gallons=("gross", "sum"),
        observed_temp=("temp", "mean"),
        api_gravity=("api", "mean"),
        terminal=("terminal", "first"),
        product=("product", "first"),
        consignee_name=("consignee_name", "first"),
    ).reset_index()

    # resolve consignee NAME → coded master (normalized match), unmapped → raw name
    chart_pairs = chart_pairs or []
    raw_to_coded = {_norm(r): c for r, c in chart_pairs}
    coded_self = {_norm(c): c for _, c in chart_pairs}

    def resolve(name):
        n = _norm(name)
        return raw_to_coded.get(n) or coded_self.get(n) or (name if name else None)

    lifts["customer_id"] = lifts["consignee_name"].map(resolve)
    lifts["lift_datetime"] = lifts["ship_date"]
    out = lifts[["customer_id", "lift_datetime", "net_gallons", "gross_gallons", "observed_temp",
                 "api_gravity", "terminal", "product", "bol_number"]].copy()

    with db.lock():
        if replace:
            db.truncate(con, schema.LIFTS)
        written = db.insert_df(con, schema.LIFTS, out)
        db.rebuild_customers_from_lifts(con, replace=replace)
        db.set_meta(con, "profile", "real_book")
        db.set_meta(con, "last_import_at", now)
        db.set_meta(con, "last_import_table", schema.LIFTS)
    n_mapped = int(lifts["customer_id"].isin(set(coded_self.values()) | set(raw_to_coded.values())).sum())
    return {
        "control_rows_dropped": n_control,
        "compartment_rows": n_compartments,
        "lifts_written": written,
        "distinct_customers": int(lifts["customer_id"].nunique()),
        "lifts_mapped_to_master": n_mapped,
        "date_min": str(lifts["ship_date"].min().date()) if written else None,
        "date_max": str(lifts["ship_date"].max().date()) if written else None,
    }


# ---- Deal book → deals table -----------------------------------------------------
_DEAL_FILES = {
    dealbook.SOURCE_TERM: "deals_summary.xlsx",
    dealbook.SOURCE_FORWARD: "forward_fixed_price_sales.xlsx",
    dealbook.SOURCE_SPOT: "wholesale_spot_deal_report.xlsx",
}


def load_deal_source(con, source: str, path: str, now: str) -> dict:
    """Parse one deal source and idempotently upsert it, then re-resolve masters via the bridge."""
    rows = dealbook.parse_source(source, path)
    for r in rows:
        r["deal_key"] = dealbook.deal_key(
            r["source"], r["customer_raw"], r["product_family"], r.get("terminal"),
            r.get("month"), r.get("deal_date"))
        r.pop("_flags", None)
    with db.lock():
        res = db.upsert_deals(con, rows, source, path.split("/")[-1], now)
        db.resolve_deal_masters(con)
        db.set_meta(con, "last_deal_import_at", now)
    res["source"] = source
    res["parsed_rows"] = len(rows)
    return res


# ---- One-shot full real-book load ------------------------------------------------
def load_real_book(con, deals_dir: str, now: str) -> dict:
    """Load the whole book in dependency order: chart → BOLs → deals."""
    import glob
    import os
    chart_path = os.path.join(deals_dir, "account_reference_chart.xlsx")
    # any *bols*.csv — so a later year's book (e.g. Jul-2024→now) just drops in and concatenates
    bol_paths = sorted(glob.glob(os.path.join(deals_dir, "*[bB][oO][lL]*.csv")))
    report: dict = {}
    chart_pairs = []
    if os.path.exists(chart_path):
        report["chart"] = load_account_chart(con, chart_path, now)
        chart_pairs = report["chart"].pop("pairs", [])
    if bol_paths:
        report["bols"] = load_bols(con, bol_paths, chart_pairs, now)
        report["bols"]["files"] = [os.path.basename(p) for p in bol_paths]
    deals: dict = {}
    for source, fname in _DEAL_FILES.items():
        p = os.path.join(deals_dir, fname)
        if os.path.exists(p):
            deals[source] = load_deal_source(con, source, p, now)
    report["deals"] = deals
    return report
