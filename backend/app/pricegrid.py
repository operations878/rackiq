"""Price-grid & landed-cost ingestion — the SELL and COST sides the margin layer reads.

Two operator workbooks feed Phase-2 margin (see docs/margin/MODELING_DECISION.md). Like the deal
book, the three sources have completely different shapes, so each gets a dedicated **format-aware
parser** (no generic mapper) that lands into one of three idempotent stores:

  • SELL grid — ``1__Wholesale_Prices___Costs_V1.xlsx``:
      - "Matrix" sheet: rows are PRODUCT+CUSTOMER concatenated with NO delimiter
        ("ULSHO4416 Oil Corp"); columns are daily dates; values are customer sell $/gal. Keys are
        split by the longest known product prefix; an unmatched key is FLAGGED, never guessed.
      - per-terminal/product sheets ("B10 ULSHO Bronx", "Baltimore ULSD", "Newark B5 ULSD"): a
        multi-row header (a weekday-number row, then a row with "Customer" + date headers, then
        customer rows of daily $/gal). Cleaner than the Matrix → PREFERRED; the Matrix only fills
        gaps (resolved at read time in margin.py).
      - "Benchmarks" sheet: named differentials/adders (DD, RACK/GEC, ASHBY) by B10/B20.
  • LANDED cost — ``6__/7__Trips_Report.xls``: barge discharge trips with per-gallon logistics legs
    (Barge / Inspector / Operational / Gain-Loss Cost Per Gallon), Estimated Trip Value, Pricing
    Type, Fixed Differential, Product Code, Product Vol (BARRELS, thousand-barrel "mb"), Discharge
    Terminal/ETA, VEF.

Stores: ``price_grid`` (sell), ``landed_costs`` (cost), ``price_differentials`` (named diffs). They
survive demo reload / reset (uploaded real data, created by :func:`ensure_tables`, not in
``schema.ALL_TABLES``) exactly like ``deals``. Customer names resolve through the SAME
``customer_crosswalk``; product codes through ``dealbook.product_family``; blend numbers are product
attributes, not identity.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import re

import pandas as pd

from . import dealbook
from .ingest import _norm, _similarity
from .margin_config import DEFAULT_CONFIG, MarginConfig

# ---- stores (created on demand; survive reset/demo like the deal book) -----------
PRICE_GRID_DDL = """CREATE TABLE IF NOT EXISTS price_grid (
    price_key VARCHAR PRIMARY KEY,
    source VARCHAR,                 -- 'matrix' | 'terminal_sheet'
    customer_raw VARCHAR,
    customer_master VARCHAR,        -- resolved coded master (NULL until the crosswalk maps it)
    product_family VARCHAR,
    product_raw VARCHAR,
    blend VARCHAR,                  -- B5 / B10 / B20 / B99 (a product attribute, not identity)
    terminal VARCHAR,               -- NULL for Matrix rows (no terminal)
    price_date DATE,
    sell_price DOUBLE,              -- customer sell price ($/gal)
    sheet VARCHAR,
    source_file VARCHAR,
    imported_at VARCHAR
)"""

LANDED_COSTS_DDL = """CREATE TABLE IF NOT EXISTS landed_costs (
    trip_key VARCHAR PRIMARY KEY,
    terminal VARCHAR,
    product_family VARCHAR,
    product_raw VARCHAR,
    discharge_date DATE,
    barge_cost DOUBLE,              -- $/gal
    inspector_cost DOUBLE,          -- $/gal
    operational_cost DOUBLE,        -- $/gal
    gainloss_cost DOUBLE,           -- $/gal
    logistics_cost DOUBLE,          -- sum of the four $/gal legs (always recoverable)
    est_trip_value DOUBLE,          -- $ (candidate all-in cargo value)
    pricing_type VARCHAR,           -- 'Monthly Average' | 'Fixed Diff'
    fixed_differential DOUBLE,      -- cargo differential ($/gal), where Pricing Type = Fixed Diff
    volume_bbl DOUBLE,              -- barrels
    volume_gal DOUBLE,              -- barrels × 42
    vol_unit VARCHAR,               -- 'bbl' | 'mb' (the magnitude heuristic's decision)
    vef DOUBLE,
    all_in_landed DOUBLE,           -- $/gal all-in (ETV/gal + logistics) IF ETV passes the sanity band; else NULL
    cost_basis VARCHAR,             -- 'all_in' | 'logistics_only'
    source_file VARCHAR,
    imported_at VARCHAR
)"""

PRICE_DIFFERENTIALS_DDL = """CREATE TABLE IF NOT EXISTS price_differentials (
    diff_key VARCHAR PRIMARY KEY,
    name VARCHAR,                   -- DD | RACK | GEC | ASHBY ...
    blend VARCHAR,                  -- B10 | B20 | NULL
    value DOUBLE,                   -- $/gal
    source_file VARCHAR,
    imported_at VARCHAR
)"""

_GRID_COLS = ["price_key", "source", "customer_raw", "customer_master", "product_family",
              "product_raw", "blend", "terminal", "price_date", "sell_price", "sheet",
              "source_file", "imported_at"]
_COST_COLS = ["trip_key", "terminal", "product_family", "product_raw", "discharge_date",
              "barge_cost", "inspector_cost", "operational_cost", "gainloss_cost", "logistics_cost",
              "est_trip_value", "pricing_type", "fixed_differential", "volume_bbl", "volume_gal",
              "vol_unit", "vef", "all_in_landed", "cost_basis", "source_file", "imported_at"]
_DIFF_COLS = ["diff_key", "name", "blend", "value", "source_file", "imported_at"]

# Known benchmark/differential names on the "Benchmarks" sheet.
_DIFF_NAMES = ("ASHBY", "RACK", "GEC", "DD")


def ensure_tables(con) -> None:
    con.execute(PRICE_GRID_DDL)
    con.execute(LANDED_COSTS_DDL)
    con.execute(PRICE_DIFFERENTIALS_DDL)


# ---- low-level workbook reading (uniform raw rows across .xlsx / .xls / .csv) ------
def _sheet_names(path: str) -> list[str]:
    if path.lower().endswith(".csv"):
        return ["csv"]
    if path.lower().endswith(".xls"):
        try:
            return list(pd.ExcelFile(path).sheet_names)
        except Exception as e:  # pragma: no cover - legacy .xls needs the xlrd engine
            raise RuntimeError(
                f"could not open '{path}' — legacy .xls needs the xlrd engine; re-save as .xlsx"
            ) from e
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _sheet_rows(path: str, sheet: str | None = None) -> list[tuple]:
    """Return a sheet's raw cell values as a list of row-tuples (dates/numbers stay typed)."""
    low = path.lower()
    if low.endswith(".csv"):
        df = pd.read_csv(path, header=None, dtype=object)
        return [tuple(r) for r in df.itertuples(index=False, name=None)]
    if low.endswith(".xls"):
        df = pd.read_excel(path, sheet_name=sheet or 0, header=None)
        return [tuple(r) for r in df.itertuples(index=False, name=None)]
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


# ---- Matrix concat-key splitting -------------------------------------------------
def split_matrix_key(key: object, cfg: MarginConfig = DEFAULT_CONFIG) -> tuple[str | None, str | None]:
    """Split a Matrix ``PRODUCT+CUSTOMER`` key (no delimiter) by the longest known product prefix.

    "ULSHO4416 Oil Corp" → ("ULSHO", "4416 Oil Corp"). A key that starts with no known product
    prefix returns (None, None) — flagged ambiguous by the caller, never guessed.
    """
    s = re.sub(r"\s+", " ", str(key).strip()) if key is not None else ""
    if not s:
        return None, None
    up = s.upper()
    # longest prefix first so "B10 ULSHO" beats "ULSHO" / a bare blend
    for pref in sorted(cfg.product_prefixes, key=len, reverse=True):
        p = pref.upper()
        if up.startswith(p):
            customer = s[len(pref):].strip(" -:")
            if customer:
                return pref, customer
    return None, None


def _date_columns(header: tuple) -> dict[int, dt.date]:
    out: dict[int, dt.date] = {}
    for j, c in enumerate(header):
        d = dealbook._to_date(c)
        if d is not None:
            out[j] = d
    return out


def _best_date_header_row(rows: list[tuple], min_dates: int = 3) -> int | None:
    """Index of the row with the most date-parseable cells (the date-header row)."""
    best, best_n = None, 0
    for i, r in enumerate(rows[:40]):
        n = len(_date_columns(r))
        if n > best_n:
            best, best_n = i, n
    return best if best_n >= min_dates else None


# ---- Matrix sheet ----------------------------------------------------------------
def parse_matrix(rows: list[tuple], cfg: MarginConfig = DEFAULT_CONFIG) -> tuple[list[dict], list[str]]:
    """Parse the Matrix sheet (concat keys × daily date columns) into price rows. Returns
    (rows, ambiguous_keys) where ambiguous_keys are concat keys that couldn't be split."""
    hdr_i = _best_date_header_row(rows)
    if hdr_i is None:
        return [], []
    date_cols = _date_columns(rows[hdr_i])
    # the key column = the first non-date column (typically col 0)
    key_col = next((j for j in range(len(rows[hdr_i])) if j not in date_cols), 0)
    out: list[dict] = []
    ambiguous: list[str] = []
    for r in rows[hdr_i + 1:]:
        if key_col >= len(r) or r[key_col] is None:
            continue
        raw_key = str(r[key_col]).strip()
        if not raw_key:
            continue
        product_raw, customer_raw = split_matrix_key(raw_key, cfg)
        if customer_raw is None:
            ambiguous.append(raw_key)
            continue
        blend = dealbook._trailing_blend(product_raw) or _leading_blend(product_raw)
        fam = dealbook.product_family(product_raw)
        cust = dealbook.base_customer_identity(customer_raw)
        for j, d in date_cols.items():
            if j >= len(r):
                continue
            px = dealbook._to_num(r[j])
            if px is None or px <= 0:
                continue
            out.append(_grid_row("matrix", cust, fam, product_raw, blend, None, d, px, "Matrix"))
    return out, ambiguous


def _leading_blend(name: str) -> str | None:
    m = re.match(r"\s*(B\d{1,3})\b", str(name or ""))
    return m.group(1) if m else None


# ---- per-terminal/product sheet --------------------------------------------------
def parse_sheet_name(name: str) -> tuple[str, str | None, str | None]:
    """Parse a terminal sheet name into (product_family, blend, terminal).

    "B10 ULSHO Bronx" → ("ULSHO","B10","Bronx"); "Baltimore ULSD" → ("ULSD",None,"Baltimore").
    """
    tokens = re.sub(r"\s+", " ", str(name).strip()).split(" ")
    blend = next((t for t in tokens if re.fullmatch(r"B\d{1,3}", t.upper())), None)
    prod_tok = None
    for t in tokens:
        if blend and t == blend:
            continue
        if dealbook.product_family(t) != "OTHER":
            prod_tok = t
            break
    fam = dealbook.product_family(prod_tok) if prod_tok else "UNKNOWN"
    rest = [t for t in tokens if t != blend and t != prod_tok]
    terminal = " ".join(rest).strip() or None
    return fam, blend, terminal


def parse_terminal_sheet(rows: list[tuple], sheet_name: str,
                         cfg: MarginConfig = DEFAULT_CONFIG) -> list[dict]:
    """Parse a per-terminal sheet (Customer column + daily date headers) into price rows."""
    fam, blend, terminal = parse_sheet_name(sheet_name)
    # find the header row carrying "Customer" (the multi-row header's second line)
    hdr_i = cust_col = None
    for i, r in enumerate(rows[:40]):
        for j, c in enumerate(r):
            if isinstance(c, str) and _norm(c) == "customer":
                hdr_i, cust_col = i, j
                break
        if hdr_i is not None:
            break
    if hdr_i is None:
        # no explicit Customer header → fall back to the best date-header row, key col 0
        hdr_i = _best_date_header_row(rows)
        cust_col = 0
        if hdr_i is None:
            return []
    date_cols = {j: d for j, d in _date_columns(rows[hdr_i]).items() if j != cust_col}
    if not date_cols:
        return []
    out: list[dict] = []
    for r in rows[hdr_i + 1:]:
        if cust_col >= len(r) or r[cust_col] is None:
            continue
        cust_raw = str(r[cust_col]).strip()
        if not cust_raw or _norm(cust_raw) in ("customer", "total", "average", "avg"):
            continue
        cust = dealbook.base_customer_identity(cust_raw)
        for j, d in date_cols.items():
            if j >= len(r):
                continue
            px = dealbook._to_num(r[j])
            if px is None or px <= 0:
                continue
            out.append(_grid_row("terminal_sheet", cust, fam, None, blend, terminal, d, px, sheet_name))
    return out


# ---- Benchmarks sheet ------------------------------------------------------------
def parse_benchmarks(rows: list[tuple]) -> list[dict]:
    """Parse the Benchmarks sheet (named differentials/adders by blend) — best-effort & tolerant."""
    out: list[dict] = []
    # find a header row carrying blend labels (B10/B20) to attribute columns
    blend_cols: dict[int, str] = {}
    for r in rows[:12]:
        for j, c in enumerate(r):
            if isinstance(c, str) and re.fullmatch(r"B\d{1,3}", c.strip().upper()):
                blend_cols[j] = c.strip().upper()
        if blend_cols:
            break
    for r in rows:
        first = next((c for c in r if c is not None and str(c).strip()), None)
        if first is None:
            continue
        nm = str(first).strip().upper()
        match = next((d for d in _DIFF_NAMES if d in nm), None)
        if match is None:
            continue
        wrote = False
        for j, c in enumerate(r):
            v = dealbook._to_num(c)
            if v is None or (isinstance(c, str) and any(d in str(c).upper() for d in _DIFF_NAMES)):
                continue
            blend = blend_cols.get(j)
            out.append({"name": match, "blend": blend, "value": v})
            wrote = True
        if not wrote:
            out.append({"name": match, "blend": None, "value": None})
    return out


# ---- Trips report (landed cost) --------------------------------------------------
_TRIP_TARGETS = {
    "discharge_date": ["discharge eta", "discharge final", "discharge date", "eta", "discharge"],
    "terminal": ["discharge terminal", "terminal", "destination"],
    "product_raw": ["product code", "product", "grade"],
    "volume_bbl": ["product vol", "volume", "barrels", "bbl", "mb", "quantity"],
    "barge_cost": ["barge cost per gallon", "barge cost", "barge"],
    "inspector_cost": ["inspector cost per gallon", "inspector cost", "inspector"],
    "operational_cost": ["operational cost per gallon", "operational cost", "operational", "ops cost"],
    "gainloss_cost": ["gain loss cost per gallon", "gain/loss cost per gallon", "gain loss cost",
                      "gainloss", "gain loss"],
    "est_trip_value": ["estimated trip value", "trip value", "est trip value"],
    "pricing_type": ["pricing type", "price type"],
    "fixed_differential": ["fixed differential", "differential", "fixed diff"],
    "vef": ["discharge final / vef", "vef", "vessel experience factor"],
}


def _match_trip_columns(header: tuple) -> dict[str, int]:
    """Fuzzy-map a Trips header row to the target fields (robust to column order / friendly names)."""
    norm = [(_norm(c) if c is not None else "") for c in header]
    out: dict[str, int] = {}
    for target, syns in _TRIP_TARGETS.items():
        best_j, best_s = None, 0.0
        for j, h in enumerate(norm):
            if not h:
                continue
            s = max((_similarity(h, _norm(syn)) for syn in syns), default=0.0)
            if s > best_s:
                best_j, best_s = j, s
        if best_j is not None and best_s >= 0.6:
            out[target] = best_j
    return out


def parse_trips(path: str, cfg: MarginConfig = DEFAULT_CONFIG) -> list[dict]:
    """Parse the barge Trips report into landed-cost rows (one per discharge trip)."""
    rows = _sheet_rows(path)
    if not rows:
        return []
    # find the header row = the one matching the most trip targets
    hdr_i, best = 0, -1
    for i, r in enumerate(rows[:30]):
        m = _match_trip_columns(r)
        if len(m) > best:
            hdr_i, best = i, len(m)
    cols = _match_trip_columns(rows[hdr_i])
    if "discharge_date" not in cols and "terminal" not in cols:
        return []
    out: list[dict] = []
    for r in rows[hdr_i + 1:]:
        def g(key):
            j = cols.get(key)
            return r[j] if j is not None and j < len(r) else None

        ddate = dealbook._to_date(g("discharge_date"))
        terminal = (str(g("terminal")).strip() if g("terminal") is not None else None) or None
        product_raw = (str(g("product_raw")).strip() if g("product_raw") is not None else None) or None
        barge = dealbook._to_num(g("barge_cost"))
        inspector = dealbook._to_num(g("inspector_cost"))
        operational = dealbook._to_num(g("operational_cost"))
        gainloss = dealbook._to_num(g("gainloss_cost"))
        legs = [x for x in (barge, inspector, operational, gainloss) if x is not None]
        if ddate is None and terminal is None and not legs:
            continue
        logistics = round(sum(legs), 6) if legs else None
        vol_raw = dealbook._to_num(g("volume_bbl"))
        vol_bbl, vol_unit = _resolve_volume(vol_raw, cfg)
        vol_gal = vol_bbl * cfg.gallons_per_barrel if vol_bbl is not None else None
        etv = dealbook._to_num(g("est_trip_value"))
        all_in, cost_basis = _all_in_landed(etv, vol_gal, logistics, cfg)
        out.append({
            "terminal": terminal,
            "product_family": dealbook.product_family(product_raw) if product_raw else "UNKNOWN",
            "product_raw": product_raw,
            "discharge_date": ddate,
            "barge_cost": barge, "inspector_cost": inspector,
            "operational_cost": operational, "gainloss_cost": gainloss,
            "logistics_cost": logistics,
            "est_trip_value": etv,
            "pricing_type": (str(g("pricing_type")).strip() if g("pricing_type") is not None else None),
            "fixed_differential": dealbook._to_num(g("fixed_differential")),
            "volume_bbl": vol_bbl, "volume_gal": vol_gal, "vol_unit": vol_unit,
            "vef": dealbook._to_num(g("vef")),
            "all_in_landed": all_in, "cost_basis": cost_basis,
        })
    return out


def _resolve_volume(raw: float | None, cfg: MarginConfig) -> tuple[float | None, str | None]:
    """Recover barrels from Trips Product Vol (labeled in barrels, written thousand-barrel 'mb')."""
    if raw is None or raw <= 0:
        return (None, None)
    if raw < cfg.mb_threshold_bbl:        # small magnitude ⇒ thousand-barrel "mb"
        return (raw * 1000.0, "mb")
    return (raw, "bbl")


def _all_in_landed(etv: float | None, vol_gal: float | None, logistics: float | None,
                   cfg: MarginConfig) -> tuple[float | None, str]:
    """All-in landed $/gal = ETV/gal + logistics, but ONLY if ETV/gal lands in the flat band
    (i.e. it embeds the index). Otherwise logistics-only (cargo flat is the un-loaded index gap)."""
    if etv is not None and vol_gal and vol_gal > 0:
        per_gal = etv / vol_gal
        if cfg.etv_flat_lo <= per_gal <= cfg.etv_flat_hi:
            return (round(per_gal + (logistics or 0.0), 6), "all_in")
    return (None, "logistics_only")


# ---- row builders + stable keys --------------------------------------------------
def _grid_row(source, customer_raw, fam, product_raw, blend, terminal, d, px, sheet) -> dict:
    return {"source": source, "customer_raw": customer_raw, "product_family": fam,
            "product_raw": product_raw, "blend": blend, "terminal": terminal,
            "price_date": d, "sell_price": round(float(px), 6), "sheet": sheet}


def price_key(source, customer_raw, fam, blend, terminal, d) -> str:
    parts = [source, _norm(customer_raw), (fam or "").lower(), (blend or "").lower(),
             (_norm(terminal) if terminal else ""), d.isoformat() if d else ""]
    return hashlib.sha1("|".join(parts).encode()).hexdigest()[:20]


def trip_key(terminal, fam, d, vol_bbl, barge) -> str:
    parts = [(_norm(terminal) if terminal else ""), (fam or "").lower(),
             d.isoformat() if d else "", f"{round(vol_bbl or 0)}", f"{round(barge or 0, 4)}"]
    return hashlib.sha1("|".join(parts).encode()).hexdigest()[:20]


def diff_key(name, blend) -> str:
    return hashlib.sha1(f"{(name or '').lower()}|{(blend or '').lower()}".encode()).hexdigest()[:16]


# ---- idempotent upserts (delete-then-insert on the stable PK) ---------------------
def _upsert(con, table: str, cols: list[str], rows: list[dict], key: str) -> int:
    if not rows:
        return 0
    agg = {r[key]: r for r in rows}          # within-file: last write wins for a repeated key
    deduped = list(agg.values())
    keys = [r[key] for r in deduped]
    ph = ", ".join("?" for _ in keys)
    con.execute(f"DELETE FROM {table} WHERE {key} IN ({ph})", keys)
    df = pd.DataFrame(deduped)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = df[cols]
    con.register("_pg_ins", df)
    try:
        sel = ", ".join(
            f'CAST("{c}" AS {_col_type(table, c)}) AS "{c}"' for c in cols)
        con.execute(f"INSERT INTO {table} ({', '.join(cols)}) SELECT {sel} FROM _pg_ins")
    finally:
        con.unregister("_pg_ins")
    return len(deduped)


def _col_type(table: str, col: str) -> str:
    if col in ("price_date", "discharge_date"):
        return "DATE"
    if col in ("sell_price", "barge_cost", "inspector_cost", "operational_cost", "gainloss_cost",
               "logistics_cost", "est_trip_value", "fixed_differential", "volume_bbl", "volume_gal",
               "vef", "all_in_landed", "value"):
        return "DOUBLE"
    return "VARCHAR"


def upsert_price_grid(con, rows: list[dict], source_file: str, now: str) -> int:
    for r in rows:
        r["price_key"] = price_key(r["source"], r["customer_raw"], r["product_family"],
                                   r.get("blend"), r.get("terminal"), r.get("price_date"))
        r["customer_master"] = None
        r["source_file"] = source_file
        r["imported_at"] = now
    return _upsert(con, "price_grid", _GRID_COLS, rows, "price_key")


def upsert_landed_costs(con, rows: list[dict], source_file: str, now: str) -> int:
    for r in rows:
        r["trip_key"] = trip_key(r.get("terminal"), r.get("product_family"),
                                 r.get("discharge_date"), r.get("volume_bbl"), r.get("barge_cost"))
        r["source_file"] = source_file
        r["imported_at"] = now
    return _upsert(con, "landed_costs", _COST_COLS, rows, "trip_key")


def upsert_differentials(con, rows: list[dict], source_file: str, now: str) -> int:
    keyed = []
    for r in rows:
        r = dict(r)
        r["diff_key"] = diff_key(r.get("name"), r.get("blend"))
        r["source_file"] = source_file
        r["imported_at"] = now
        keyed.append(r)
    return _upsert(con, "price_differentials", _DIFF_COLS, keyed, "diff_key")


# ---- master resolution (raw grid name → confirmed crosswalk master) ---------------
def resolve_price_masters(con) -> int:
    """Set ``price_grid.customer_master`` from the confirmed crosswalk (raw grid name → master).

    Mirrors ``db.resolve_deal_masters``: unmapped names stay NULL (no fabricated master). Idempotent.
    """
    con.execute("UPDATE price_grid SET customer_master = NULL")
    con.execute("""
        UPDATE price_grid SET customer_master = cw.master_id
        FROM customer_crosswalk cw
        WHERE TRIM(price_grid.customer_raw) = cw.variant_key
          AND cw.status = 'confirmed' AND cw.master_id IS NOT NULL
    """)
    return int(con.execute(
        "SELECT count(*) FROM price_grid WHERE customer_master IS NOT NULL").fetchone()[0])


def unmapped_grid_customers(con, limit: int = 500) -> list[dict]:
    """Raw grid customer names that don't yet resolve to a confirmed crosswalk master."""
    rows = con.execute("""
        SELECT customer_raw, count(*) AS n, count(DISTINCT price_date) AS days
        FROM price_grid WHERE customer_master IS NULL
        GROUP BY 1 ORDER BY n DESC LIMIT ?""", [limit]).fetchall()
    return [{"customer_raw": r[0], "rows": int(r[1]), "days": int(r[2])} for r in rows]


# ---- reads -----------------------------------------------------------------------
def read_price_grid(con) -> pd.DataFrame:
    ensure_tables(con)
    return con.execute(f"SELECT {', '.join(_GRID_COLS)} FROM price_grid").df()


def read_landed_costs(con) -> pd.DataFrame:
    ensure_tables(con)
    return con.execute(f"SELECT {', '.join(_COST_COLS)} FROM landed_costs").df()


def read_differentials(con) -> pd.DataFrame:
    ensure_tables(con)
    return con.execute(f"SELECT {', '.join(_DIFF_COLS)} FROM price_differentials").df()


def store_counts(con) -> dict:
    ensure_tables(con)
    return {
        "price_grid_rows": int(con.execute("SELECT count(*) FROM price_grid").fetchone()[0]),
        "price_grid_customers": int(con.execute(
            "SELECT count(DISTINCT customer_raw) FROM price_grid").fetchone()[0]),
        "price_grid_mapped": int(con.execute(
            "SELECT count(DISTINCT customer_master) FROM price_grid "
            "WHERE customer_master IS NOT NULL").fetchone()[0]),
        "landed_cost_trips": int(con.execute("SELECT count(*) FROM landed_costs").fetchone()[0]),
        "landed_cost_all_in": int(con.execute(
            "SELECT count(*) FROM landed_costs WHERE cost_basis = 'all_in'").fetchone()[0]),
        "differentials": int(con.execute("SELECT count(*) FROM price_differentials").fetchone()[0]),
    }


# ---- workbook-level parse + load (the repeatable, idempotent ingestion) -----------
def parse_price_workbook(path: str, cfg: MarginConfig = DEFAULT_CONFIG) -> dict:
    """Parse the whole wholesale-price workbook: Matrix + per-terminal sheets + Benchmarks."""
    prices: list[dict] = []
    differentials: list[dict] = []
    ambiguous: list[str] = []
    sheets_parsed: list[str] = []
    for sn in _sheet_names(path):
        low = sn.strip().lower()
        rows = _sheet_rows(path, sn)
        if not rows:
            continue
        if low == "matrix":
            p, amb = parse_matrix(rows, cfg)
            prices += p
            ambiguous += amb
            sheets_parsed.append(sn)
        elif "benchmark" in low:
            differentials += parse_benchmarks(rows)
            sheets_parsed.append(sn)
        else:
            p = parse_terminal_sheet(rows, sn, cfg)
            if p:
                prices += p
                sheets_parsed.append(sn)
    return {"prices": prices, "differentials": differentials,
            "ambiguous_keys": ambiguous, "sheets_parsed": sheets_parsed}


def load_price_grid_file(con, path: str, now: str, cfg: MarginConfig = DEFAULT_CONFIG) -> dict:
    """Parse + idempotently upsert the wholesale-price workbook, then resolve masters."""
    ensure_tables(con)
    parsed = parse_price_workbook(path, cfg)
    fname = path.split("/")[-1]
    n_prices = upsert_price_grid(con, parsed["prices"], fname, now)
    n_diffs = upsert_differentials(con, parsed["differentials"], fname, now)
    mapped = resolve_price_masters(con)
    return {"prices_written": n_prices, "differentials_written": n_diffs,
            "masters_resolved": mapped, "ambiguous_keys": parsed["ambiguous_keys"],
            "sheets_parsed": parsed["sheets_parsed"], "filename": fname}


def load_trips_file(con, path: str, now: str, cfg: MarginConfig = DEFAULT_CONFIG) -> dict:
    """Parse + idempotently upsert the barge Trips report into ``landed_costs``."""
    ensure_tables(con)
    rows = parse_trips(path, cfg)
    fname = path.split("/")[-1]
    n = upsert_landed_costs(con, rows, fname, now)
    all_in = sum(1 for r in rows if r.get("cost_basis") == "all_in")
    return {"trips_written": n, "all_in_trips": all_in,
            "logistics_only_trips": n - all_in, "filename": fname}


# File-name patterns for the one-shot loader.
def _is_price_workbook(name: str) -> bool:
    n = name.lower()
    return ("wholesale" in n and "price" in n) or "prices_costs" in n or n.startswith("1__")


def _is_trips_file(name: str) -> bool:
    return "trip" in name.lower()


def load_price_book(con, directory: str, now: str, cfg: MarginConfig = DEFAULT_CONFIG) -> dict:
    """One-shot: load the wholesale price workbook + the Trips report from a directory."""
    import glob
    import os
    ensure_tables(con)
    report: dict = {}
    for p in sorted(glob.glob(os.path.join(directory, "*"))):
        base = os.path.basename(p)
        if _is_price_workbook(base) and base.lower().endswith((".xlsx", ".xlsm")):
            report["prices"] = load_price_grid_file(con, p, now, cfg)
        elif _is_trips_file(base) and base.lower().endswith((".xls", ".xlsx", ".csv")):
            report.setdefault("trips", []).append(load_trips_file(con, p, now, cfg))
    return report
