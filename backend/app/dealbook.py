"""Deal book ingestion — term, forward-fixed, and spot deals → the canonical ``deals`` table.

The deal book is the **commitment spine**: it records what volume each customer is *contracted*
to take (term + forward-fixed) and what they buy *opportunistically* (spot). The variability
score is computed from BOL lifts alone; the deal book only **annotates** it with commitment
context ("82% term-backed", "spot-only", "requirements"). It is also the spine that later phases
(margin, position/hedging) read.

The three sources have three completely different shapes, so there is **no generic column
mapper** — each gets a dedicated, format-aware parser that lands rows into one canonical schema:

    master customer × product family × terminal × month, tagged term|forward_fixed|spot,
    with committed gallons (term/forward) or realized gallons + price (spot).

Layouts (locked during discovery):
  • TERM  — ``deals_summary.xlsx`` sheet "Term": a month-pivot blocked by product. Col A carries
            the product (forward-filled: "2 OIL", "4 Oil", "Diesel", "B99"), col B the customer.
            Each customer is a *volume row* immediately followed by a *basis-price row*. A
            month-header row (cols ≥ C = month names) precedes each block; the schedule is
            month-only (no year) so the year is inferred (``TERM_BASE_YEAR``) and flagged. A
            trailing "NET"/"gross" token marks the volume basis; a "requirements" cell marks a
            requirements (no-fixed-number) contract.
  • FORWARD-FIXED — ``forward_fixed_price_sales.xlsx`` sheet "Active Deals": a month pivot whose
            month headers (real dates 2020→2027) live in row index 1, cols 5..85. Col A is EITHER
            a customer name OR the status "Approved" — the customer forward-fills down a section.
            Col D = deal date, col E = locked $/gal, cols 5..85 = committed gallons per month.
            Orphan rows above the first customer name are excluded; the REMAINING column is ignored
            (its values are price-like, not gallons). Customer/product/terminal may appear in cols
            A/B/C.
  • SPOT  — ``wholesale_spot_deal_report.xlsx``: one sheet per month, clean columns
            (Date, Representative, Company, Product, Gallons, Price). Company = customer; the row's
            own Date sets the month (robust to a deal logged on the "wrong" tab).
"""

from __future__ import annotations

import datetime as dt
import hashlib
import re

from openpyxl import load_workbook

from . import db

# Term schedule is month-only (no year on the sheet). We anchor month 0 ("November") to November
# of this year and FLAG it; the commitment annotation uses annual *rates*, so it is robust to the
# exact anchor. Override via load_deals(term_base_year=...).
TERM_BASE_YEAR = 2023

SOURCE_TERM = "term"
SOURCE_FORWARD = "forward_fixed"
SOURCE_SPOT = "spot"

_MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July", "August",
     "September", "October", "November", "December"], start=1)}


# ---- Product family normalization (shared by BOL + deal ingestion) ---------------
# Canonical families. Heating oil (ULSHO) is the weather-sensitive distillate; dyed clear diesel is
# its own family; #4 is the heavy residual grade; RD/bio is renewable/biodiesel. Blend numbers
# (B5/B10/B20/B99) are a sub-attribute of the *base* family, not a separate family.
FAMILIES = ("ULSD", "ULSHO", "DYED", "HO4", "RD", "GAS", "OTHER")


def product_family(raw: object) -> str:
    """Normalize a raw product label (BOL product name OR deal product code) to a canonical family.

    Order matters: renewable/bio and #4 are checked first; heating-oil (incl. dyed HO, which is
    still heating oil) before dyed *diesel*; a bare "#2"/"2 OIL" (no diesel token) reads as heating
    oil. Returns "OTHER" for control/unknown codes (e.g. ``ZZZ``).
    """
    if raw is None:
        return "OTHER"
    s = re.sub(r"\s+", " ", str(raw).strip().upper())
    if not s or s in ("ZZZ", "NONE"):
        return "OTHER"
    has_diesel = bool(re.search(r"\bULSD\b|\bLSD\b|ULTRA L|U:LSD|DIESEL", s))
    # 1) renewable / biodiesel (the *fuel* is renewable, regardless of a "diesel" token)
    if re.search(r"RENEWABLE|BIO ?DIESEL|\bB-?99|\bR-?99|\bRD-?99|\bRD\b", s):
        return "RD"
    # 2) #4 heavy residual
    if re.search(r"#\s*4|FUEL OIL #?4|FUEL OIL 4|NO\.? ?4 OIL|\b4 OIL\b", s):
        return "HO4"
    # 3) gasoline (explicitly excluded from heating logic downstream)
    if re.search(r"\bRBOB\b|GASOLINE|\bUNL\b|\bREG\b|\bPREM\b|ETHANOL|\bE10\b|\bE15\b", s) and not has_diesel:
        return "GAS"
    # 4) heating oil (ULSHO) — incl. dyed heating oil (still heating oil) and bare #2/2-OIL
    if re.search(r"ULSHO|U:SHO|HEATING OIL|\bHO\b|\bHO[- ]?\d|\bH/?O\b", s):
        return "ULSHO"
    if re.search(r"\b#?2 ?OIL\b|\b#2\b|\b2 OIL\b", s) and not has_diesel:
        return "ULSHO"
    # 5) dyed clear diesel (off-road) — dyed + a diesel token, and NOT heating oil
    if "DYED" in s and has_diesel:
        return "DYED"
    # 6) clear ULSD
    if has_diesel:
        return "ULSD"
    # 7) leftover bare #2 / 2 oil → heating oil
    if re.search(r"\b#?2\b|\b2 OIL\b", s):
        return "ULSHO"
    return "OTHER"


# Product families that are heating fuels (weather-sensitive). The variability score exposes a
# weather-adjustment SEAM only for these; everything else is never weather-normalized.
HEATING_FAMILIES = frozenset({"ULSHO", "HO4"})


def _norm_name(s: object) -> str:
    return re.sub(r"\s+", " ", str(s).strip()) if s is not None else ""


def base_customer_identity(name: str) -> str:
    """Strip a trailing blend number so a *blend level* never splits customer identity.

    "GEC 10" / "GEC 20" → "GEC" (the 10/20 is the B10/B20 product, captured separately). Other
    names pass through unchanged.
    """
    return re.sub(r"\s+(?:B?\d{1,3})$", "", _norm_name(name)).strip()


def _trailing_blend(name: str) -> str | None:
    m = re.search(r"\s+(B?\d{1,3})$", _norm_name(name))
    return m.group(1) if m else None


def deal_key(source: str, customer_raw: str, product_family_: str, terminal: str | None,
             month: dt.date | None, deal_date: dt.date | None) -> str:
    """Stable idempotency key: customer × product × terminal × month × source × deal_date."""
    parts = [source, _norm_name(customer_raw).lower(), (product_family_ or "").lower(),
             _norm_name(terminal).lower() if terminal else "",
             month.isoformat() if month else "", deal_date.isoformat() if deal_date else ""]
    return hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:20]


def _to_num(v: object) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = re.sub(r"[,$%\s]", "", str(v).strip().replace("−", "-"))
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(v: object) -> dt.date | None:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


# ---- TERM parser (deals_summary.xlsx · "Term") ----------------------------------
def parse_term(path: str, base_year: int = TERM_BASE_YEAR) -> list[dict]:
    """Parse the term month-pivot into canonical deal rows (one per customer×product×month)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Term" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Term"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    out: list[dict] = []
    col_month: dict[int, dt.date] = {}   # column index → month date (from the current header row)
    cur_product = None
    flags = {"year_inferred": True}

    def is_month_header(r) -> bool:
        # a header row: col0/col1 blank, and ≥3 month names in cols 2+
        if _norm_name(r[0]) or _norm_name(r[1]):
            return False
        hits = sum(1 for c in r[2:] if isinstance(c, str) and c.strip().lower() in _MONTHS)
        return hits >= 3

    def build_month_map(r):
        col_month.clear()
        # The schedule runs forward from the first month; roll the year over each January.
        year = base_year
        prev = None
        for j in range(2, len(r)):
            c = r[j]
            if isinstance(c, str) and c.strip().lower() in _MONTHS:
                mo = _MONTHS[c.strip().lower()]
                if prev is not None and mo < prev:   # wrapped past December → next year
                    year += 1
                col_month[j] = dt.date(year, mo, 1)
                prev = mo

    i = 0
    n = len(rows)
    while i < n:
        r = rows[i]
        if is_month_header(r):
            build_month_map(r)
            i += 1
            continue
        c0 = _norm_name(r[0])
        c1 = _norm_name(r[1])
        # product label in col0 (carries down); ignore the NET/gross/totals tokens
        if c0 and c0.lower() not in ("net", "gross"):
            cur_product = c0
        # a customer volume row: col1 is a name and there is numeric/keyword volume in the month cols
        if c1 and col_month:
            vol_cells = {j: r[j] for j in col_month if j < len(r) and r[j] is not None}
            has_req = any(isinstance(v, str) and "require" in v.lower() for v in vol_cells.values())
            # A real volume row carries LARGE numbers (hundreds+). A basis-only row (e.g. a customer
            # whose deal is a pure differential schedule) carries tiny values (< ~1 $/gal) and must
            # NOT be read as committed gallons — require a volume-magnitude cell to qualify.
            big = any((n := _to_num(v)) is not None and abs(n) >= 10 for v in vol_cells.values())
            if vol_cells and (has_req or big) and not all(
                    isinstance(v, str) and v.strip().lower() in _MONTHS for v in vol_cells.values()):
                # volume basis token (NET / gross) sits in a trailing cell
                basis = "net"
                for c in r[2:]:
                    if isinstance(c, str) and c.strip().lower() in ("net", "gross"):
                        basis = c.strip().lower()
                # the price (basis differential) row is the NEXT row
                price_row = rows[i + 1] if i + 1 < n else ()
                customer = base_customer_identity(c1)
                blend = _trailing_blend(c1)
                fam = product_family((cur_product or "") + (" B" + re.sub(r"\D", "", blend) if blend else ""))
                if cur_product:
                    fam = product_family(_compose_term_product(cur_product, blend))
                for j, v in vol_cells.items():
                    is_req = isinstance(v, str) and "require" in v.lower()
                    gal = None if is_req else _to_num(v)
                    if not is_req and (gal is None or gal == 0):
                        continue
                    price = _to_num(price_row[j]) if j < len(price_row) else None
                    out.append({
                        "source": SOURCE_TERM,
                        "customer_raw": customer,
                        "product_raw": _compose_term_product(cur_product, blend),
                        "product_family": fam,
                        "terminal": None,
                        "month": col_month[j],
                        "committed_gallons": gal,
                        "realized_gallons": None,
                        "price": price,
                        "price_type": "basis",
                        "commitment_type": "requirements" if is_req else "firm",
                        "volume_basis": basis,
                        "deal_date": None,
                        "representative": None,
                    })
                i += 2   # consumed the volume row + its price row
                continue
        i += 1
    for row in out:
        row["_flags"] = flags
    return out


def _compose_term_product(product: str | None, blend: str | None) -> str:
    p = product or ""
    if blend:
        return f"{p} {blend}".strip()
    return p


# ---- FORWARD-FIXED parser (forward_fixed_price_sales.xlsx · "Active Deals") -------
def parse_forward_fixed(path: str) -> list[dict]:
    """Parse the forward-fixed 'Active Deals' month pivot into canonical deal rows.

    Months from the row-1 header (real dates). Customer forward-fills down col A (a row whose col A
    is the status 'Approved' or a date inherits the customer above). Orphan deals above the first
    named customer are excluded. All months with volume are stored (the reconciliation window is
    chosen later); the REMAINING column is ignored.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = "Active Deals" if "Active Deals" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        return []

    header = rows[1]
    col_month: dict[int, dt.date] = {}
    for j, c in enumerate(header):
        d = _to_date(c)
        if d is not None and j >= 4:          # the month axis starts after the ID columns
            col_month[j] = d.replace(day=1)
    if not col_month:
        return []
    month_cols = sorted(col_month)
    DEAL_DATE_COL, PRICE_COL = 3, 4

    out: list[dict] = []
    cur_customer = None
    for i in range(2, len(rows)):
        r = rows[i]
        c0 = r[0]
        if isinstance(c0, str) and c0.strip() and c0.strip().lower() != "approved":
            cur_customer = _norm_name(c0)
        if cur_customer is None:               # orphan rows above the first named customer
            continue
        deal_dt = _to_date(r[DEAL_DATE_COL]) if len(r) > DEAL_DATE_COL else None
        price = _to_num(r[PRICE_COL]) if len(r) > PRICE_COL else None
        prod_raw = _norm_name(r[1]) if len(r) > 1 and isinstance(r[1], str) else None
        terminal = _norm_name(r[2]) if len(r) > 2 and isinstance(r[2], str) else None
        fam = product_family(prod_raw) if prod_raw else "UNKNOWN"
        customer = base_customer_identity(cur_customer)
        for j in month_cols:
            gal = _to_num(r[j]) if j < len(r) else None
            if gal is None or gal == 0:
                continue
            out.append({
                "source": SOURCE_FORWARD,
                "customer_raw": customer,
                "product_raw": prod_raw,
                "product_family": fam,
                "terminal": terminal,
                "month": col_month[j],
                "committed_gallons": gal,
                "realized_gallons": None,
                "price": price,
                "price_type": "fixed",
                "commitment_type": "firm",
                "volume_basis": "unknown",
                "deal_date": deal_dt,
                "representative": None,
            })
    return out


# ---- SPOT parser (wholesale_spot_deal_report.xlsx · one sheet per month) ----------
def parse_spot(path: str) -> list[dict]:
    """Parse the spot report (one tab per month) into canonical realized-deal rows."""
    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # header row carries Company / Product / Gallons / Price (with stray whitespace)
        header = [(_norm_name(c).lower() if c is not None else "") for c in rows[0]]
        idx = {}
        for j, h in enumerate(header):
            if h.startswith("date") or h.startswith("spot deal"):  # the date column is titled "SPOT DEALS"
                idx.setdefault("date", j)
            elif h.startswith("repres"):
                idx.setdefault("rep", j)
            elif h.startswith("compan"):
                idx.setdefault("company", j)
            elif h.startswith("product"):
                idx.setdefault("product", j)
            elif h.startswith("gallon"):
                idx.setdefault("gallons", j)
            elif h.startswith("price"):
                idx.setdefault("price", j)
        if "company" not in idx or "gallons" not in idx:
            continue
        if "date" not in idx:
            # fall back to the column with the most date-like values (spot reports lead with Date)
            best, best_n = None, 0
            for j in range(min(3, max((len(r) for r in rows[1:30]), default=0))):
                n = sum(1 for r in rows[1:60] if j < len(r) and _to_date(r[j]) is not None)
                if n > best_n:
                    best, best_n = j, n
            if best is not None:
                idx["date"] = best
        for r in rows[1:]:
            company = _norm_name(r[idx["company"]]) if idx["company"] < len(r) else ""
            gal = _to_num(r[idx["gallons"]]) if idx["gallons"] < len(r) else None
            if not company or gal is None or gal == 0:
                continue
            d = _to_date(r[idx["date"]]) if "date" in idx and idx["date"] < len(r) else None
            prod_raw = _norm_name(r[idx["product"]]) if "product" in idx and idx["product"] < len(r) else None
            price = _to_num(r[idx["price"]]) if "price" in idx and idx["price"] < len(r) else None
            rep = _norm_name(r[idx["rep"]]) if "rep" in idx and idx["rep"] < len(r) else None
            month = d.replace(day=1) if d else None
            out.append({
                "source": SOURCE_SPOT,
                "customer_raw": base_customer_identity(company),
                "product_raw": prod_raw,
                "product_family": product_family(prod_raw),
                "terminal": None,
                "month": month,
                "committed_gallons": None,
                "realized_gallons": gal,
                "price": price,
                "price_type": "realized",
                "commitment_type": "firm",
                "volume_basis": "net",
                "deal_date": d,
                "representative": rep,
            })
    wb.close()
    return out


PARSERS = {SOURCE_TERM: parse_term, SOURCE_FORWARD: parse_forward_fixed, SOURCE_SPOT: parse_spot}


def parse_source(source: str, path: str) -> list[dict]:
    fn = PARSERS.get(source)
    if fn is None:
        raise ValueError(f"unknown deal source '{source}' (expected one of {list(PARSERS)})")
    return fn(path)


def detect_deal_source(path: str) -> str | None:
    """Infer which deal source a workbook is, from its sheet names (so the Data Studio 'Deals'
    upload doesn't need the operator to pick term/forward/spot by hand)."""
    try:
        wb = load_workbook(path, read_only=True)
        sheets = [s.lower() for s in wb.sheetnames]
        wb.close()
    except Exception:
        return None
    if any("active deal" in s for s in sheets):
        return SOURCE_FORWARD
    if any(s == "term" for s in sheets) or any(s == "fixed" for s in sheets):
        return SOURCE_TERM
    month_like = sum(1 for s in sheets if any(m in s for m in
                     ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]))
    if month_like >= 2:
        return SOURCE_SPOT
    return None


# ---- Crosswalk bridge: deal-book customer → BOL master ---------------------------
BRIDGE_THRESHOLD = 0.84


def bridge_candidates(con, threshold: float = BRIDGE_THRESHOLD) -> dict:
    """Stage the deal-book → BOL-master bridge (THE JOIN). Never auto-merges.

    For every distinct deal-book customer, report whether it already resolves to a master that ALSO
    appears in the BOL book (``mapped``), a best fuzzy candidate among BOL masters to confirm
    (``candidate``), or nothing close (``unmapped``). Also reports the match RATE — what share of
    committed deal volume bridges to a real BOL customer — so a thin bridge surfaces loudly before any
    annotation is trusted.
    """
    from .ingest import _norm, _similarity

    bol_masters = [r[0] for r in con.execute(
        "SELECT DISTINCT customer_id FROM lifts WHERE customer_id IS NOT NULL").fetchall()]
    bol_master_set = set(bol_masters)
    bol_norm = {_norm(m): m for m in bol_masters}

    rows = con.execute("""
        SELECT customer_raw, customer_master,
               sum(coalesce(committed_gallons, 0)) AS committed_gal,
               sum(coalesce(realized_gallons, 0)) AS realized_gal,
               count(*) AS n
        FROM deals GROUP BY 1, 2""").df()

    mapped, candidates, unmapped = [], [], []
    committed_total = float(rows["committed_gal"].sum()) or 1.0
    committed_bridged = 0.0
    for r in rows.itertuples():
        raw = r.customer_raw
        master = r.customer_master
        committed = float(r.committed_gal or 0)
        entry = {"customer_raw": raw, "committed_gallons": round(committed, 0),
                 "realized_gallons": round(float(r.realized_gal or 0), 0), "deal_rows": int(r.n)}
        if master and master in bol_master_set:
            entry["master"] = master
            mapped.append(entry)
            committed_bridged += committed
            continue
        # propose the best BOL master by fuzzy similarity on the base identity
        n = _norm(base_customer_identity(raw))
        best, best_s = None, 0.0
        for mn, m in bol_norm.items():
            s = _similarity(n, mn)
            if s > best_s:
                best, best_s = m, s
        entry["candidate_master"] = best
        entry["similarity"] = round(best_s, 3)
        (candidates if best_s >= threshold else unmapped).append(entry)

    candidates.sort(key=lambda e: -e["committed_gallons"])
    unmapped.sort(key=lambda e: -e["committed_gallons"])
    mapped.sort(key=lambda e: -e["committed_gallons"])
    return {
        "n_deal_customers": int(len(rows)),
        "n_mapped": len(mapped), "n_candidates": len(candidates), "n_unmapped": len(unmapped),
        "match_rate_by_committed_volume": round(100 * committed_bridged / committed_total, 1),
        "match_rate_by_count": round(100 * len(mapped) / max(1, len(rows)), 1),
        "mapped": mapped, "candidates": candidates, "unmapped": unmapped,
        "threshold": threshold,
    }


def confirm_bridge(con, pairs: list[tuple[str, str]], now: str) -> dict:
    """Confirm staged deal-name → BOL-master bridges (writes the crosswalk, re-resolves deals).

    ``pairs`` is ``[(deal_raw_name, bol_master), ...]`` — each becomes a confirmed crosswalk entry so
    the deal customer rolls up to that BOL master, then ``deals.customer_master`` is re-resolved.
    """
    entries = []
    for raw, master in pairs:
        raw, master = str(raw).strip(), str(master).strip()
        if not raw or not master:
            continue
        entries.append({"variant_key": raw, "master_id": master, "master_name": master,
                        "confidence": 1.0, "status": "confirmed", "source": "deal_bridge",
                        "updated_at": now})
    n = db.upsert_crosswalk_entries(con, entries)
    resolved = db.resolve_deal_masters(con)
    return {"confirmed": len(entries), "crosswalk_written": n, "deal_rows_resolved": resolved}
